from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from vosk import Model, KaldiRecognizer
import pandas as pd
import openpyxl
import pyaudio
import os
from common_functions import clear_root_buttons

class Entries:
    def __init__(self, root, acc, second_frame, i, opt, ind, add = False, snow = 1):
        self.root = root
        self.acc = acc
        self.row_num = i
        self.second_frame = second_frame
        self.opt = opt
        self.index = ind
        self.df_acc = None
        self.lock = False
        self.snow = snow
        self.wid = {7: 50, 13: 17}
        clear_root_buttons(root)
        model = Model(os.getcwd()+"\\vosk-model-small-en-us-0.15")
        self.recognizer = KaldiRecognizer(model, 16000)
        if (os.path.exists('Report.xlsx')):
            book = openpyxl.load_workbook('Report.xlsx')
            sheets = {ws.title: ws for ws in book.worksheets}
            if self.acc in sheets:
                self.df_acc = pd.read_excel('Report.xlsx', sheet_name=self.acc)
        if (opt == 'insert'):
            #fetch SNOW
            self.create_row(0, 4, 15,"")
            if (os.path.exists('Report.xlsx') and not add):
                if self.acc in sheets:
                    self.snow = int(self.df_acc.iloc[-1]['SNOW']) + 1
                    self.create_row(4, 5, 15,self.snow)
                else:
                    self.create_row(4, 5, 15,self.snow)
            else:
                self.create_row(4, 5, 15,self.snow)

            self.create_row(5, 7, 15,"")
            self.create_row(7, 8, 50,"")
            self.create_row(8, 13, 15,"")
            self.create_row(13, 14, 17,"")
        else:
            for j in range(1, 15):
                mywid = 15 if (j-1 not in self.wid) else self.wid[j-1]
                cell = Text(self.second_frame, height=3, width=mywid)
                cell.grid(row=2, column=j-1)
                sen = "" if pd.isna(self.df_acc.iloc[self.index['orow']][j]) else self.df_acc.iloc[self.index['orow']][j]
                cell.insert(END, sen)

        ttk.Button(self.root, text="RECORD",
                   command=self.record).place(x=200, y=600)
        ttk.Button(self.root, text="STOP RECORD",
                   command=self.stoprec).place(x=295, y=600)
        ttk.Button(self.root, text="SAVE", command=lambda w=second_frame:
                   self.get_all_entry_widgets_text_content(w)).place(x=400, y=600)
        if (opt == 'insert'):
            ttk.Button(self.root, text="ADD ROW",
                    command=self.add_row).place(x=500, y=600)
        elif(opt == 'update'):
            ttk.Button(self.root, text="CANCEL",
                    command=self.cancel).place(x=500, y=600)

    def create_row(self, scol, ecol, wid, ins):
        for j in range(scol, ecol):
            self.cell = Text(self.second_frame, height=3, width=wid)
            self.cell.grid(row=self.row_num, column=j)
            if ins:
                self.cell.insert(INSERT, ins)
            self.cell.bind("<Double-Button-1>", self.OnTextClick)

    def OnTextClick(self, event):
        if self.lock:
            print("Listening...")
            self.textdata = ""
            mic = pyaudio.PyAudio()
            stream = mic.open(format=pyaudio.paInt16, channels=1,
                              rate=16000, input=True, frames_per_buffer=8192)
            stream.start_stream()
            bol = True
            while bol:
                data = stream.read(4096)
                if self.recognizer.AcceptWaveform(data):
                    self.textdata = self.recognizer.Result()
                    if len(self.textdata) > 17:
                        print(self.textdata)
                        self.second_frame.focus_get().insert(
                            INSERT, self.textdata[14:-3]+" ")
                    else:
                        print("No speech recognised.")
                    bol = False

    def get_all_entry_widgets_text_content(self, parent_widget):
        if self.opt == 'insert':
            children_widgets = parent_widget.winfo_children()
            self.mydata = [child_widget.get(1.0, 'end')[:-1] for child_widget in children_widgets if child_widget.winfo_class() == 'Text']
            self.mdata = [[self.acc]+self.mydata[i:i+14] for i in range(0, len(self.mydata), 14)]

            df = pd.DataFrame(self.mdata, columns=["Account Number", "Time & Date", "Airframe Hours", "How found or Aircrew Code", "By Whom (Name)", "SNOW", "Reason for placing unserviceable", "Job Number", "Repairs, Storage instructions, Modifications, Special Technical Instructions Servicing Instructions or other work including component Replacements. Note 1. State serial Numbers of Replacement items,  where applicable 2. RN only. Quote Workshop Reference ORN",
                            "Signature of Operative(s)", "Time and Date Completed", "Trade", "Man Hours", "Signature of Supervisor", "Authorised signature Certified Defect Cleared or Transferred to MOD from 703/704"])
            if not (os.path.exists('Report.xlsx')):
                df.to_excel('Report.xlsx', index=False, sheet_name=self.acc)
                print("EXCEL CREATED SUCCESSFULLY")
                ttk.Label(self.root, text="EXCEL CREATED SUCCESSFULLY", foreground='green').place(x=600, y=600)
            else:
                book = openpyxl.load_workbook('Report.xlsx')
                writer = pd.ExcelWriter('Report.xlsx', engine='openpyxl')
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                if self.acc not in writer.sheets:
                    df.to_excel('Report.xlsx', index=False, sheet_name=self.acc)
                    df.to_excel(writer, sheet_name=self.acc, startrow=0, index=False)
                else:
                    sr = writer.sheets[self.acc].max_row
                    df.to_excel(writer, sheet_name=self.acc, startrow=sr, index=False, header=False)
                writer.save()
                print("EXCEL UPDATED SUCCESSFULLY")
                messagebox.showinfo('Saved Changes', "EXCEL UPDATED SUCCESSFULLY")
            from scrollableframe import ScrollableFrame
            ScrollableFrame(self.root, 'insert', self.acc)
        elif self.opt == 'update':
            children_widgets = parent_widget.winfo_children()
            self.mydata = [child_widget.get(1.0, 'end')[:-1] for child_widget in children_widgets if child_widget.winfo_class() == 'Text']
            wb = openpyxl.load_workbook('Report.xlsx')
            wb.get_sheet_names()
            sheet1 = wb.get_sheet_by_name(self.acc)
            for x in range(2,16):
                cell = sheet1.cell(row=self.index['orow']+2, column=x)
                cell.value = self.mydata[x-2]
            wb.save('Report.xlsx')
            print("UPDATED SUCCESSFULLY")
            messagebox.showinfo('Saved Changes', "EXCEL UPDATED SUCCESSFULLY")


    def record(self):
        self.lock = True
        print("Record BUTTON CLICKED")

    def stoprec(self):
        self.lock = False
        print("STOP Record BUTTON CLICKED")

    def add_row(self):
        self.row_num = self.row_num + 1
        self.snow = self.snow + 1
        Entries(self.root, self.acc, self.second_frame, self.row_num, 'insert', None, True, self.snow)
    
    def cancel(self):
        self.root.destroy()