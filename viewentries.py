from tkinter import *
from tkinter import ttk
from common_functions import clear_root_buttons, getnum
import pandas as pd
import openpyxl

class ViewEntries:
    def __init__(self, root, acc, second_frame, row_no):
        self.row_no = row_no
        self.second_frame = second_frame
        self.root = root
        self.row_range = 0
        self.acc = acc
        self.limit = 3
        self.wid = {7: 50, 13: 17}
        clear_root_buttons(root)
        book = openpyxl.load_workbook('Report.xlsx')
        self.sheets = {ws.title: ws for ws in book.worksheets}
        if self.acc in self.sheets:
            self.df_acc = pd.read_excel('Report.xlsx', sheet_name=self.acc)
            self.mrows = self.sheets[self.acc].max_row - 1
            self.mcols = self.sheets[self.acc].max_column
            for r in range(self.row_no, self.limit+self.row_no):
                b = ttk.Button(self.second_frame, text="Update", command= self.update)
                b.grid(row=r, column=0, ipady=10.5, ipadx=5)
            self.my_display(0)
        else:
            ttk.Label(self.second_frame, text="NO RECORDS FOUND!!!", borderwidth=3, font=("Courier", 30)).grid(row=self.row_no, column=0, columnspan=8, pady=100)
            

    def focus(self, event):
        widget = self.second_frame.focus_get()
        print(widget, "has focus")
        print(event)
    
    def update(self):
        widget = self.second_frame.focus_get()
        # print("=======>", widget)
        vrow = int(getnum(str(widget)[-8:]))
        index = {'orow': vrow+self.row_range-3, 'vrow' : vrow}
        print(index)
        top= Toplevel(self.root)
        top.geometry("900x600")
        top.title("UPDATE ENTRY")
        from scrollableframe import ScrollableFrame
        ScrollableFrame(top, 'update', self.acc, index)

    def my_display(self, offset):
        for i in range(self.row_no, self.limit+self.row_no):
            for j in range(1, self.mcols):
                mywid = 15 if (j-1 not in self.wid) else self.wid[j-1]
                cell = Text(self.second_frame, height=3, width=mywid)
                cell.grid(row=i, column=j)

                sen = ""
                if (i-self.row_no+offset>=self.mrows):
                    cell.insert(END, "")
                else:
                    sen = "" if pd.isna(self.df_acc.iloc[i-self.row_no+offset][j]) else self.df_acc.iloc[i-self.row_no+offset][j]
                    cell.insert(END, sen)
                self.row_range = i-self.row_no+offset
                
        back = offset - self.limit
        next = offset + self.limit 
        b1 = ttk.Button(self.root, text='Next >', command=lambda: self.my_display(next))
        b1.place(x=200,y=600)
        b2 = ttk.Button(self.root, text='< Prev', command=lambda: self.my_display(back))
        b2.place(x=350,y=600)
        if(self.mrows <= next): 
            b1["state"]="disabled"
        else:
            b1["state"]="active"  
            
        if(back >= 0):
            b2["state"]="active"  
        else:
            b2["state"]="disabled"
        
        # print(self.row_range)

